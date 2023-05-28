import datetime
import sqlite3
import sys
import os
from os import system
import pymsgbox as amsg
import openpyxl
from openpyxl import load_workbook
from PyQt5 import QtGui
from PyQt5 import QtCore
from PyQt5 import QtWidgets
from PyQt5 import uic
from PyQt5.QtWidgets import QTableWidgetItem, QAbstractItemView, QDateEdit, QTimeEdit, QDateTimeEdit, QApplication, QMainWindow, QMenu
#from uavto_ui import Ui_MainWindow
import locale
locale.setlocale(locale.LC_ALL, '')
'ru_RU.utf8'
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.getcwd()
    return os.path.join(base_path, relative_path)
with sqlite3.connect("utrbase.db") as db:
     cur = db.cursor()
db.commit()
msg = amsg.confirm
class MyWin(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.error_dialog = QtWidgets.QErrorMessage()
        #self.ui = Ui_MainWindow() # Экземпляр класса Ui_MainWindow, в нем конструктор всего GUI.
        #self.ui.setupUi(self)  # Инициализация GUI
        self.uavto_ui = uic.loadUi(resource_path("uavto_ui.ui"))
        self.uavto_ui.show()
        self.uavto_ui.action_8.triggered.connect(self.unit_2)
        self.uavto_ui.action_9.triggered.connect(sys.exit)
        self.uavto_ui.action_7.triggered.connect(self.unit_13)
        self.uavto_ui.action_2.triggered.connect(self.unit_3)
        self.uavto_ui.action_3.triggered.connect(self.unit_4)
        self.uavto_ui.action_6.triggered.connect(self.unit15)
        self.uavto_ui.action_5.triggered.connect(self.unit11)
        self.uavto_ui.action.triggered.connect(self.unit_5)
        # Открыть новую форму
    def unit11(self):
        self.dialog11 = uic.loadUi(resource_path("unit11.ui"))
        self.dialog11.show()
        self.dialog11.pushButton.clicked.connect(self.unit11_TRAN)
        self.dialog11.pushButton_2.clicked.connect(self.run_unit11)
        self.dialog11.pushButton_3.clicked.connect(self.dialog11.close)
        self.dialog11.dateEdit.setDate(datetime.datetime.now())
        self.dialog11.dateEdit_2.setDate(datetime.datetime.now())
    def run_unit11(self):
        wb = load_workbook(f"{os.getcwd()}/kart.xltx")
        ws1 = wb['Лист1']
        nach= self.dialog11.dateEdit.dateTime().toString('yyyy.MM.dd ddd')
        conc = self.dialog11.dateEdit_2.dateTime().toString('yyyy.MM.dd ddd')
        nach_1 = self.dialog11.dateEdit.dateTime().toString('dd.MM.yyyy')
        conc_1 = self.dialog11.dateEdit_2.dateTime().toString('dd.MM.yyyy')
        Nt = self.dialog11.label_2.text()
        Gosn = self.dialog11.label_3.text()
        ws1['C2'].value = Nt
        ws1['F6'].value = Gosn
        ws1['C8'].value = nach_1
        ws1['F8'].value = conc_1
        transport = self.dialog11.label_2.text()
        row = [i[0] for i in db.execute(f""" SELECT ID FROM listi WHERE (nt ='{transport}') and (Data>='{nach}') and (Data<='{conc}') ORDER BY Data""").fetchall()]
        stroka = 1
        for row in row:
        #row = db.execute(f""" SELECT ID FROM listi WHERE (nt ='{transport}') and (Data>='{nach}') and (Data<='{conc}') ORDER BY Data""").fetchall()
            Data = db.execute(""" SELECT Data From listi WHERE id == ?""", (row,)).fetchone()[0]
            #Adrp = (db.execute(""" SELECT adrp From listi WHERE id == ?""", (row,)).fetchone()[0])
            Numlist = db.execute(""" SELECT Numlist From listi WHERE id == ?""", (row,)).fetchone()[0]
            #Nt = (db.execute(""" SELECT Nt From listi WHERE id == ?""", (row,)).fetchone()[0])
            #Gosn = (db.execute(""" SELECT Gosn From listi WHERE id == ?""", (row,)).fetchone()[0])
            #Vrn = (db.execute(""" SELECT Vrn From listi WHERE id == ?""", (row,)).fetchone()[0])
            #Vrv = (db.execute(""" SELECT Vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
            Ovozv = float(db.execute(""" SELECT Ovozv From listi WHERE id == ?""", (row,)).fetchone()[0])
            #Psn = (db.execute(""" SELECT Psn From listi WHERE id == ?""", (row,)).fetchone()[0])
            Oviezd = float(db.execute(""" SELECT Oviezd From listi WHERE id == ?""", (row,)).fetchone()[0])
            Doz = float(db.execute(""" SELECT Doz From listi WHERE id == ?""", (row,)).fetchone()[0])
            Rnorm = float(db.execute(""" SELECT Rnorm From listi WHERE id == ?""", (row,)).fetchone()[0])
            Rfact = float(db.execute(""" SELECT Rfact From listi WHERE id == ?""", (row,)).fetchone()[0])
            #Econ = (db.execute(""" SELECT Econ From listi WHERE id == ?""", (row,)).fetchone()[0])
            #Pereras = (db.execute(""" SELECT Pereras From listi WHERE id == ?""", (row,)).fetchone()[0])
            Psv = float(db.execute(""" SELECT Psv From listi WHERE id == ?""", (row,)).fetchone()[0])
            # час наряда self.dialog_5_RED.lineEdit_27.setText(db.execute(""" SELECT vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
            proy = float(db.execute(""" SELECT proy From listi WHERE id == ?""", (row,)).fetchone()[0])
            Fam = (db.execute(""" SELECT Fam From listi WHERE id == ?""", (row,)).fetchone()[0])
            Ima = (db.execute(""" SELECT Ima From listi WHERE id == ?""", (row,)).fetchone()[0])
            Otch = (db.execute(""" SELECT Otch From listi WHERE id == ?""", (row,)).fetchone()[0])
            #Ud = (db.execute(""" SELECT Ud From listi WHERE id == ?""", (row,)).fetchone()[0])
            #klass = (db.execute(""" SELECT klass From listi WHERE id == ?""", (row,)).fetchone()[0])
            #ws1['A' + str(stroka + 13)].value = stroka
            #ws1['B' + str(stroka + 13)].value = Nt
            ws1['B' + str(stroka + 12)].value = ((Fam)+' '+(Ima)+' '+(Otch))
            ws1['A' + str(stroka + 12)].value = Data
            ws1['C' + str(stroka + 12)].value = Numlist
            if proy != '':
                ws1['D' + str(stroka + 12)].value = proy
            if Oviezd != '':
                ws1['E' + str(stroka + 12)].value = Oviezd
            if Doz != '':
                ws1['F' + str(stroka + 12)].value = Doz
            if Rnorm != '':
                ws1['G' + str(stroka + 12)].value = Rfact
            if Ovozv != '':
                ws1['H' + str(stroka + 12)].value = Ovozv
            stroka = stroka + 1
        stroka_2 = stroka + 1
        ws1['A' + str(stroka_2 + 12)].value = 'ИТОГО'
        try:
          if proy != '':
             ws1['D' + str(stroka_2 + 12)] = f"=SUM(D13:D{str(stroka + 12)}"
             ws1['D' + str(stroka_2 + 12)].value
          if Doz != '':
             ws1['F' + str(stroka_2 + 12)] = f"=SUM(F13:F{str(stroka + 12)}"
             ws1['F' + str(stroka_2 + 12)].value
          if Rnorm != '':
             ws1['G' + str(stroka_2 + 12)] = f"=SUM(G13:G{str(stroka + 12)}"
             ws1['G' + str(stroka_2 + 12)].value
          if Ovozv != '':
             ws1['H' + str(stroka_2 + 12)] = f"=(H{str((stroka-1) + 12)})"
             ws1['H' + str(stroka_2 + 12)].value
        except:
          msg('ДЛЯ ВЫБАННОГО ТРАНСПОРТА НЕТ ПУТЕВОК', 'НА ВЫБРАННЫЙ ПЕРИОД', ['ПРОДОЛЖТЬ'])
        wb.save("kart1.xlsx")
        wb.close()
        system('xdg-open ' + "kart1.xlsx")


    def unit11_TRAN(self):
        self.dialog11_TRAN = uic.loadUi(resource_path("unit_5_NEW_TRAN.ui"))
        self.dialog11_TRAN.show()
        self.dialog11_TRAN.pushButton_2.clicked.connect(self.dialog11_TRAN.close)
        self.dialog11_TRAN.pushButton.clicked.connect(self.unit11_TRAN_1)
        cur.execute('''SELECT nt, gn FROM Tran''')
        self.dialog11_TRAN.tableWidget.setRowCount(0)
        for row, form in enumerate(cur):
            self.dialog11_TRAN.tableWidget.insertRow(row)
            for column, item in enumerate(form):
                self.dialog11_TRAN.tableWidget.setItem(row, column, QTableWidgetItem(str(item)))
    def unit11_TRAN_1(self):
        i = self.dialog11_TRAN.tableWidget.currentRow()
        nt = self.dialog11_TRAN.tableWidget.item(i, 0).text()
        gn = self.dialog11_TRAN.tableWidget.item(i, 1).text()
        self.dialog11.label_2.setText(str(nt))
        self.dialog11.label_3.setText(str(gn))
        self.dialog11_TRAN.close()

        #     speed = db.execute("""SELECT speed FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn,)).fetchone()[0]
        #     ost = db.execute("""SELECT ost FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn,)).fetchone()[0]
        #     norml = db.execute("""SELECT norml FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn,)).fetchone()[0]
        #     normz = db.execute("""SELECT normz FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn,)).fetchone()[0]
         #self.dialog11.label_2.setText(str(nt))
         #self.dialog11.label_3.setText(str(gn))
        #     self.dialog11.lineEdit_2.setText(str(nt))
        #     self.dialog11.lineEdit_3.setText(str(gn))
        #     self.dialog11.lineEdit_16.setText(str('%g' % round((float(speed)), 1)))
        #     self.dialog11.lineEdit_21.setText(str('%g' % round((float(ost)), 1)))
        #     if int(db.execute(""" SELECT norml From sett""").fetchone()[0]) == 1:
        #         self.dialog_5_NEW.lineEdit_22.setText(str('%g' % round((float(norml)), 1)))
        #     else:
        #         self.dialog_5_NEW.lineEdit_22.setText(str('%g' % round((float(normz)), 1)))
        #     self.dialog_5_NEW_TRAN.close()




    def unit15(self):
        self.dialog15 = uic.loadUi(resource_path("unit15.ui"))
        self.dialog15.show()
        self.dialog15.dateEdit.setDate(datetime.datetime.now())
        self.dialog15.dateEdit_2.setDate(datetime.datetime.now())
        self.dialog15.pushButton.clicked.connect(self.unit15_run)
        self.dialog15.pushButton_2.clicked.connect(self.dialog15.close)
    def unit15_run(self):
        wb = load_workbook("reestr.xltx")
        ws1 = wb['Лист1']
        nach = self.dialog15.dateEdit.dateTime().toString('yyyy.MM.dd ddd')
        conc = self.dialog15.dateEdit_2.dateTime().toString('yyyy.MM.dd ddd')
        row = [i[0] for i in db.execute(
            f""" SELECT ID FROM listi WHERE (Data>='{nach}') and (Data<='{conc}') ORDER BY Data""").fetchall()]
        stroka = 1
        for row in row:
            # row = db.execute(f""" SELECT ID FROM listi WHERE (nt ='{transport}') and (Data>='{nach}') and (Data<='{conc}') ORDER BY Data""").fetchall()
            Data = (db.execute(""" SELECT Data From listi WHERE id == ?""", (row,)).fetchone()[0])
            Adrp = (db.execute(""" SELECT adrp From listi WHERE id == ?""", (row,)).fetchone()[0])
            Numlist = (str(db.execute(""" SELECT Numlist From listi WHERE id == ?""", (row,)).fetchone()[0]))
            Nt = (db.execute(""" SELECT Nt From listi WHERE id == ?""", (row,)).fetchone()[0])
            Gosn = (db.execute(""" SELECT Gosn From listi WHERE id == ?""", (row,)).fetchone()[0])
            Vrn = (db.execute(""" SELECT Vrn From listi WHERE id == ?""", (row,)).fetchone()[0])
            Vrv = (db.execute(""" SELECT Vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
            Ovozv = (db.execute(""" SELECT Ovozv From listi WHERE id == ?""", (row,)).fetchone()[0])
            Psn = (db.execute(""" SELECT Psn From listi WHERE id == ?""", (row,)).fetchone()[0])
            Oviezd = (db.execute(""" SELECT Oviezd From listi WHERE id == ?""", (row,)).fetchone()[0])
            Doz = (db.execute(""" SELECT Doz From listi WHERE id == ?""", (row,)).fetchone()[0])
            Rnorm = (db.execute(""" SELECT Rnorm From listi WHERE id == ?""", (row,)).fetchone()[0])
            Rfact = (db.execute(""" SELECT Rfact From listi WHERE id == ?""", (row,)).fetchone()[0])
            Econ = (db.execute(""" SELECT Econ From listi WHERE id == ?""", (row,)).fetchone()[0])
            Pereras = (db.execute(""" SELECT Pereras From listi WHERE id == ?""", (row,)).fetchone()[0])
            Psv = (db.execute(""" SELECT Psv From listi WHERE id == ?""", (row,)).fetchone()[0])
            # час наряда self.dialog_5_RED.lineEdit_27.setText(db.execute(""" SELECT vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
            proy = (db.execute(""" SELECT proy From listi WHERE id == ?""", (row,)).fetchone()[0])
            Fam = (db.execute(""" SELECT Fam From listi WHERE id == ?""", (row,)).fetchone()[0])
            Ima = (db.execute(""" SELECT Ima From listi WHERE id == ?""", (row,)).fetchone()[0])
            Otch = (db.execute(""" SELECT Otch From listi WHERE id == ?""", (row,)).fetchone()[0])
            Ud = (db.execute(""" SELECT Ud From listi WHERE id == ?""", (row,)).fetchone()[0])
            klass = (db.execute(""" SELECT klass From listi WHERE id == ?""", (row,)).fetchone()[0])
            ws1['A' + str(stroka + 3)].value = stroka
            ws1['B' + str(stroka + 3)].value = Nt
            ws1['C' + str(stroka + 3)].value = ((Fam) + ' ' + (Ima) + ' ' + (Otch))
            ws1['D' + str(stroka + 3)].value = Data
            ws1['E' + str(stroka + 3)].value = Numlist
            ws1['F' + str(stroka + 3)].value = Adrp
            ws1['G' + str(stroka + 3)].value = Vrn
            ws1['H' + str(stroka + 3)].value = Vrv
            if Psn != '':
                ws1['I' + str(stroka + 3)].value = Psn
            if Psv != '':
                ws1['J' + str(stroka + 3)].value = Psv
            if proy != '':
                ws1['K' + str(stroka + 3)].value = proy
            if Oviezd != '':
                ws1['L' + str(stroka + 3)].value = Oviezd
            if Doz != '':
                ws1['M' + str(stroka + 3)].value = Doz
            if Rnorm != '':
                ws1['N' + str(stroka + 3)].value = Rfact
            if Ovozv != '':
                ws1['O' + str(stroka + 3)].value = Ovozv
            stroka = stroka + 1
        wb.save("reestr1.xlsx")
        wb.close()
        system('xdg-open ' + "reestr1.xlsx")


    def unit_2(self):
        self.dialog_2 = uic.loadUi(resource_path("unit_2.ui"))
        self.dialog_2.show()
        self.dialog_2.pushButton.clicked.connect(self.dialog_2.close)
    def unit_13(self):
        self.dialog_13 = uic.loadUi(resource_path("unit_13.ui"))
        ListNum = [i[0] for i in db.execute("""select Numlist from listi""").fetchall()]
        intNum = [x for x in ListNum if isinstance(x, (int, float))]
        rowNum = max(intNum)
        db.execute("""UPDATE sett set num == ?""", (rowNum,))
        self.dialog_13.lineEdit.setText(str(db.execute(""" SELECT num From sett""").fetchone()[0]))
        self.dialog_13.lineEdit_8.setText(str(db.execute(""" SELECT login From sett""").fetchone()[0]))
        self.dialog_13.lineEdit_9.setText(str(db.execute(""" SELECT pass From sett""").fetchone()[0]))
        self.dialog_13.lineEdit_2.setText(str(db.execute(""" SELECT org From sett""").fetchone()[0]))
        self.dialog_13.lineEdit_3.setText(str(db.execute(""" SELECT adr From sett""").fetchone()[0]))
        self.dialog_13.lineEdit_4.setText(str(db.execute(""" SELECT tel From sett""").fetchone()[0]))
        self.dialog_13.lineEdit_5.setText(str(db.execute(""" SELECT mex From sett""").fetchone()[0]))
        self.dialog_13.lineEdit_6.setText(str(db.execute(""" SELECT disp From sett""").fetchone()[0]))
        self.dialog_13.lineEdit_7.setText(str(db.execute(""" SELECT meds From sett""").fetchone()[0]))
        self.Radio_1()
        self.dialog_13.radioButton.toggled.connect(self.onClicked)
        self.dialog_13.show()
        self.dialog_13.pushButton_2.clicked.connect(self.dialog_13.close)
        self.dialog_13.pushButton.clicked.connect(self.pechat_line_1)
    def Radio_1(self):
        if int(db.execute(""" SELECT norml From sett""").fetchone()[0]) == 1:
            self.dialog_13.radioButton.setChecked(True)
        else:
            self.dialog_13.radioButton_2.setChecked(True)
    def onClicked(self):
        if self.dialog_13.radioButton.isChecked():
            db.execute("""Update sett set norml == ? WHERE id == 1""", (1,))
        else:
            db.execute("""Update sett set norml == ? WHERE id == 1""", (0,))
    def readPass(self):
        pass
    def pechat_line_1(self):
        db.execute("""Update sett set num == ? WHERE id == 1""", ([self.dialog_13.lineEdit.text()]))
        db.execute("""Update sett set login == ? WHERE id == 1""", ([self.dialog_13.lineEdit_8.text()]))
        db.execute("""Update sett set pass == ? WHERE id == 1""", ([self.dialog_13.lineEdit_9.text()]))
        db.execute("""Update sett set org == ? WHERE id == 1""", ([self.dialog_13.lineEdit_2.text()]))
        db.execute("""Update sett set adr == ? WHERE id == 1""", ([self.dialog_13.lineEdit_3.text()]))
        db.execute("""Update sett set tel == ? WHERE id == 1""", ([self.dialog_13.lineEdit_4.text()]))
        db.execute("""Update sett set mex == ? WHERE id == 1""", ([self.dialog_13.lineEdit_5.text()]))
        db.execute("""Update sett set disp == ? WHERE id == 1""", ([self.dialog_13.lineEdit_6.text()]))
        db.execute("""Update sett set meds == ? WHERE id == 1""", ([self.dialog_13.lineEdit_7.text()]))
        db.commit()
        self.dialog_13.close()
    def unit_3(self):
        self.uavto_ui.stackedWidget.setCurrentIndex(0)
        self.dialog_3 = uic.loadUi(resource_path("unit_3.ui"))
        self.uavto_ui.gridLayout.addWidget(self.dialog_3)
        self.dialog_3.show()
        self.uavto_ui.horizontalLayout_3.removeWidget(self.dialog_3)
        self.dialog_3.action_11.triggered.connect(self.unit_6)
        self.dialog_3.action_12.triggered.connect(self.unit_6_1)
        self.dialog_3.action_13.triggered.connect(self.unit_6_confirm)
        self.dialog_3.action_14.triggered.connect(self.CLOSES_dialog_3)
        self.v_vod_3()
    def CLOSES_dialog_3(self):
        self.uavto_ui.stackedWidget.setCurrentIndex(2)
        self.dialog_3.close()
        self.v_vod_3()
    def v_vod_3(self):
        cur.execute('''SELECT fam, ima, Otch, Ud, klass FROM vod''')
        self.dialog_3.tableWidget.setRowCount(0)
        for row, form in enumerate(cur):
            self.dialog_3.tableWidget.insertRow(row)
            for column, item in enumerate(form):
                self.dialog_3.tableWidget.setItem(row, column, QTableWidgetItem(str(item)))
    def unit_6(self):
        self.dialog_6 = uic.loadUi(resource_path("unit_6.ui"))
        self.dialog_6.show()
        self.dialog_6.PushButton.clicked.connect(self.newline)
        self.dialog_6.PushButton_2.clicked.connect(self.dialog_6.close)
    def newline(self):
        fam = self.dialog_6.lineEdit.text()
        ima = self.dialog_6.lineEdit_2.text()
        Otch = self.dialog_6.lineEdit_3.text()
        Ud = self.dialog_6.lineEdit_4.text()
        klass = self.dialog_6.lineEdit_5.text()
        db.execute("""Insert into vod ( Fam, ima, Otch, Ud, klass) values (?,?,?,?,?)""", (
         fam, ima, Otch, Ud, klass ))
        db.commit()
        self.v_vod_3()
        self.dialog_6.close()
    def unit_6_1(self):
        try:
          row = self.NomID(self.x)
          self.dialog_6_1 = uic.loadUi(resource_path("unit_6.ui"))
          self.dialog_6_1.show()
          self.dialog_6_1.PushButton.clicked.connect(self.RED)
          self.dialog_6_1.PushButton_2.clicked.connect(self.dialog_6_1.close)
          self.dialog_6_1.lineEdit.setText(str(db.execute(""" SELECT Fam From vod WHERE id == ?""", (row, )).fetchone()[0]))
          self.dialog_6_1.lineEdit_2.setText(str(db.execute(""" SELECT ima From vod WHERE id == ?""", (row, )).fetchone()[0]))
          self.dialog_6_1.lineEdit_3.setText(str(db.execute(""" SELECT Otch From vod WHERE id == ?""", (row, )).fetchone()[0]))
          self.dialog_6_1.lineEdit_4.setText(str(db.execute(""" SELECT Ud From vod WHERE id == ?""", (row, )).fetchone()[0]))
          self.dialog_6_1.lineEdit_5.setText(str(db.execute(""" SELECT Klass From vod WHERE id == ?""", (row, )).fetchone()[0]))
        except:
          msg('Нужно выбрать строку', 'ВНИМАНИЕ', ['ВЫБРАТЬ'])
           #self.dialog_6_1.close()
    def RED(self):
        row = self.NomID(self.x)
        db.execute("""Update vod set Fam == ? WHERE id == ?""", ((self.dialog_6_1.lineEdit.text()), row, ))
        db.execute("""Update vod set ima == ? WHERE id == ?""", ((self.dialog_6_1.lineEdit_2.text()), row, ))
        db.execute("""Update vod set Otch == ? WHERE id == ?""", ((self.dialog_6_1.lineEdit_3.text()), row, ))
        db.execute("""Update vod set Ud == ? WHERE id == ?""", ((self.dialog_6_1.lineEdit_4.text()), row, ))
        db.execute("""Update vod set Klass == ? WHERE id == ?""", ((self.dialog_6_1.lineEdit_5.text()), row, ))
        db.commit()
        self.v_vod_3()
        self.dialog_6_1.close()
    def unit_6_confirm(self):
        try:
          row = self.NomID(self.x)
          self.dialog_6_confirm = uic.loadUi(resource_path("unit_6_confirm.ui"))
          self.dialog_6_confirm.show()
          self.dialog_6_confirm.pushButton.clicked.connect(self.Del)
          self.dialog_6_confirm.pushButton_2.clicked.connect(self.dialog_6_confirm.close)
        except:
          msg('Нужно выбрать строку', 'ВНИМАНИЕ', ['ВЫБРАТЬ'])
           #self.dialog_6_confirm.close()
    def Del(self):
        row = self.NomID(self.x)
        db.execute("""DELETE FROM vod WHERE id == ?""", (row,))
        db.commit()
        self.v_vod_3()
        self.dialog_6_confirm.close()
    def NomID(self, x):
        i = self.dialog_3.tableWidget.currentRow()
        Fam = self.dialog_3.tableWidget.item(i, 0).text()
        ima = self.dialog_3.tableWidget.item(i, 1).text()
        Otch = self.dialog_3.tableWidget.item(i, 2).text()
        Ud = self.dialog_3.tableWidget.item(i, 3).text()
        klass = self.dialog_3.tableWidget.item(i, 4).text()
        x = db.execute("""SELECT id FROM vod WHERE (Fam, ima, Otch, Ud, klass) == (?, ?, ?, ?, ?)""", (
        Fam, ima, Otch, Ud, klass,)).fetchone()[0]
        return x
    def unit_4(self):
        self.uavto_ui.stackedWidget.setCurrentIndex(1)
        self.dialog_4 = uic.loadUi(resource_path("unit_4.ui"))
        self.uavto_ui.gridLayout_2.addWidget(self.dialog_4)
        self.dialog_4.show()
        self.uavto_ui.horizontalLayout_4.removeWidget(self.dialog_4)
        self.dialog_4.action_5.triggered.connect(self.CLOSES_dialog_4)
        self.dialog_4.action.triggered.connect(self.unit_4_NEW)
        self.dialog_4.action_2.triggered.connect(self.unit_4_RED)
        self.dialog_4.action_3.triggered.connect(self.unit_4_DEL)
        # self.dialog_4.action_3.triggered.connect(self.unit_заполнитьссервера)
        self.v_vod_4()
    def CLOSES_dialog_4(self):
        self.uavto_ui.stackedWidget.setCurrentIndex(2)
        self.dialog_4.close()
        self.v_vod_4()
    def v_vod_4(self):
        cur.execute('''SELECT nt, gn, speed, ost, numt, regn, nomr, norml, normz FROM Tran''')
        self.dialog_4.tableWidget.setRowCount(0)
        for row, form in enumerate(cur):
            self.dialog_4.tableWidget.insertRow(row)
            for column, item in enumerate(form):
                self.dialog_4.tableWidget.setItem(row, column, QTableWidgetItem(str(item)))
    def unit_4_NEW(self):
        self.dialog_4_NEW = uic.loadUi(resource_path("unit_4_NEW.ui"))
        self.dialog_4_NEW.show()
        self.dialog_4_NEW.lineEdit_3.setText('0')
        self.dialog_4_NEW.lineEdit_4.setText('0')
        self.dialog_4_NEW.lineEdit_8.setText('0')
        self.dialog_4_NEW.lineEdit_9.setText('0')
        self.dialog_4_NEW.pushButton.clicked.connect(self.newline_4)
        self.dialog_4_NEW.pushButton_2.clicked.connect(self.dialog_4_NEW.close)
    def newline_4(self):
        nt = self.dialog_4_NEW.lineEdit.text()
        gn = self.dialog_4_NEW.lineEdit_2.text()
        speed = self.dialog_4_NEW.lineEdit_3.text()
        ost = self.dialog_4_NEW.lineEdit_4.text()
        numt = self.dialog_4_NEW.lineEdit_5.text()
        regn = self.dialog_4_NEW.lineEdit_6.text()
        norm = self.dialog_4_NEW.lineEdit_7.text()
        norml = self.dialog_4_NEW.lineEdit_8.text()
        normz = self.dialog_4_NEW.lineEdit_9.text()
        db.execute("""Insert into Tran ( nt, gn, speed, ost, numt, regn, nomr, norml, normz) values (?,?,?,?,?,?,?,?,?)""", ( nt, numt, speed, ost, gn, regn, norm, norml, normz))
        db.commit()
        self.v_vod_4()
        self.dialog_4_NEW.close()
    def unit_4_RED(self):
        try:
            row = self.NomID_4(self.x)
            self.dialog_4_RED = uic.loadUi(resource_path("unit_4_RED.ui"))
            self.dialog_4_RED.show()
            self.dialog_4_RED.pushButton.clicked.connect(self.RED_4)
            self.dialog_4_RED.pushButton_2.clicked.connect(self.dialog_4_RED.close)
            self.dialog_4_RED.lineEdit.setText(str(db.execute(""" SELECT nt From Tran WHERE id == ?""", (row,)).fetchone()[0]))
            self.dialog_4_RED.lineEdit_2.setText(str(db.execute(""" SELECT gn From Tran WHERE id == ?""", (row,)).fetchone()[0]))
            self.dialog_4_RED.lineEdit_3.setText(str(db.execute(""" SELECT speed From Tran WHERE id == ?""", (row,)).fetchone()[0]))
            self.dialog_4_RED.lineEdit_4.setText(str(db.execute(""" SELECT ost From Tran WHERE id == ?""", (row,)).fetchone()[0]))
            self.dialog_4_RED.lineEdit_5.setText(str(db.execute(""" SELECT numt From Tran WHERE id == ?""", (row,)).fetchone()[0]))
            self.dialog_4_RED.lineEdit_6.setText(str(db.execute(""" SELECT regn From Tran WHERE id == ?""", (row,)).fetchone()[0]))
            self.dialog_4_RED.lineEdit_7.setText(str(db.execute(""" SELECT nomr From Tran WHERE id == ?""", (row,)).fetchone()[0]))
            self.dialog_4_RED.lineEdit_8.setText(str(db.execute(""" SELECT norml From Tran WHERE id == ?""", (row,)).fetchone()[0]))
            self.dialog_4_RED.lineEdit_9.setText(str(db.execute(""" SELECT normz From Tran WHERE id == ?""", (row,)).fetchone()[0]))
        except:
            msg('Нужно выбрать строку', 'ВНИМАНИЕ', ['ВЫБРАТЬ'])
            #self.dialog_4_RED.close()

    def RED_4(self):
        row = self.NomID_4(self.x)
        db.execute("""Update Tran set nt == ? WHERE id == ?""", ((self.dialog_4_RED.lineEdit.text()), row,))
        db.execute("""Update Tran set gn == ? WHERE id == ?""", ((self.dialog_4_RED.lineEdit_2.text()), row,))
        db.execute("""Update Tran set speed == ? WHERE id == ?""", ((self.dialog_4_RED.lineEdit_3.text()), row,))
        db.execute("""Update Tran set ost == ? WHERE id == ?""", ((self.dialog_4_RED.lineEdit_4.text()), row,))
        db.execute("""Update Tran set numt == ? WHERE id == ?""", ((self.dialog_4_RED.lineEdit_5.text()), row,))
        db.execute("""Update Tran set regn == ? WHERE id == ?""", ((self.dialog_4_RED.lineEdit_6.text()), row,))
        db.execute("""Update Tran set nomr == ? WHERE id == ?""", ((self.dialog_4_RED.lineEdit_7.text()), row,))
        db.execute("""Update Tran set norml == ? WHERE id == ?""", ((self.dialog_4_RED.lineEdit_8.text()), row,))
        db.execute("""Update Tran set normz == ? WHERE id == ?""", ((self.dialog_4_RED.lineEdit_9.text()), row,))
        db.commit()
        self.v_vod_4()
        self.dialog_4_RED.close()
    def NomID_4(self, x):
        i = self.dialog_4.tableWidget.currentRow()
        nt = self.dialog_4.tableWidget.item(i, 0).text()
        gn = self.dialog_4.tableWidget.item(i, 1).text()
        speed = self.dialog_4.tableWidget.item(i, 2).text()
        ost = self.dialog_4.tableWidget.item(i, 3).text()
        numt = self.dialog_4.tableWidget.item(i, 4).text()
        regn = self.dialog_4.tableWidget.item(i, 5).text()
        norm = self.dialog_4.tableWidget.item(i, 6).text()
        norml = self.dialog_4.tableWidget.item(i, 7).text()
        normz = self.dialog_4.tableWidget.item(i, 8).text()
        x = db.execute("""SELECT id FROM Tran WHERE (nt, gn, speed, ost, numt, regn, nomr, norml, normz) == (?, ?, ?, ?, ?, ?, ?, ?, ?)""", (
            nt, gn, speed, ost, numt, regn, norm, norml, normz)).fetchone()[0]
        return x
    def unit_4_DEL(self):
        try:
            row = self.NomID_4(self.x)
            self.dialog_4_DEL = uic.loadUi(resource_path("unit_4_confirm.ui"))
            self.dialog_4_DEL.show()
            self.dialog_4_DEL.pushButton.clicked.connect(self.Del_4)
            self.dialog_4_DEL.pushButton_2.clicked.connect(self.dialog_4_DEL.close)
        except:
            msg('Нужно выбрать строку', 'ВНИМАНИЕ', ['ВЫБРАТЬ'])
            #self.dialog_4_DEL.close()
    def Del_4(self):
        row = self.NomID_4(self.x)
        db.execute("""DELETE FROM Tran WHERE id == ?""", (row,))
        db.commit()
        self.v_vod_4()
        self.dialog_4_DEL.close()
    def unit_5(self):
        self.uavto_ui.stackedWidget.setCurrentIndex(2)
        self.dialog_5 = uic.loadUi(resource_path("unit_5.ui"))
        self.uavto_ui.gridLayout_3.addWidget(self.dialog_5)
        self.dialog_5.show()
        self.dialog_5.tableWidget.doubleClicked.connect(self.unit_5_RED)
        self.uavto_ui.horizontalLayout_5.removeWidget(self.dialog_5)
        self.dialog_5.action_5.triggered.connect(self.CLOSES_dialog_5)
        self.dialog_5.action.triggered.connect(self.unit_5_NEW)
        self.dialog_5.action_2.triggered.connect(self.unit_5_RED)
        self.dialog_5.action_3.triggered.connect(self.unit_5_DEL)
        menu = QMenu()
        #menu.show()
        self.dialog_5.action_4.setMenu(menu)
        menu.addAction("Легковая", self.legtr)
        menu.addAction("Грузовая", self.grutr)
        menu.addAction("Автобус").triggered.connect(self.autr)
        #self.dialog_5.action_4.setMenu(menu)
        #self.dialog_5.action_4.triggered.connect(self.unit_5_PECHAT)
        self.v_vod_5()
    #self.dialog_5.action_4.triggered.connect(self.actionClicked)
    #def actionClicked(self, action):
        #print('Action: ', action)
    def CLOSES_dialog_5(self):
        self.uavto_ui.stackedWidget.setCurrentIndex(0)
        self.dialog_5.close()
        self.v_vod_5()
    def v_vod_5(self):
        cur.execute('''SELECT Data, Numlist, Nt, Gosn, Fam, Ima, Otch, Ud, Klass, Psn, Adrp, Vrn, Vrv, Doz, Oviezd,
                     Ovozv, Rnorm, Rfact, Econ, Pereras, Psv, proy FROM listi''')
        self.dialog_5.tableWidget.setRowCount(0)
        for row, form in enumerate(cur):
            self.dialog_5.tableWidget.insertRow(row)
            for column, item in enumerate(form):
                self.dialog_5.tableWidget.setItem(row, column, QTableWidgetItem(str(item)))
    def unit_5_NEW(self):
        self.dialog_5_NEW = uic.loadUi(resource_path("unit_5_NEW.ui"))
        self.dialog_5_NEW.show()
        self.dialog_5_NEW.pushButton_6.clicked.connect(self.newline_5_NEW)
        self.dialog_5_NEW.pushButton.clicked.connect(self.unit_5_NEW_TRAN)
        self.dialog_5_NEW.pushButton_5.clicked.connect(self.unit_5_NEW_VOD)
        self.dialog_5_NEW.pushButton_7.clicked.connect(self.dialog_5_NEW.close)
        self.dialog_5_NEW.pushButton_4.clicked.connect(self.otpr)
        self.dialog_5_NEW.pushButton_2.clicked.connect(self.vr)
        self.but()
        self.e_23_26()
        self.e_17()
        self.dialog_5_NEW.dateEdit_2.setDate(datetime.datetime.now())
        self.dialog_5_NEW.dateEdit_3.setDate(datetime.datetime.now())
        self.dialog_5_NEW.dateEdit.setDate(datetime.datetime.now())
        #rowID = db.execute("""select ID from listi order by ID desc""").fetchone()[0]
        ListNum = [i[0] for i in db.execute("""select Numlist from listi""").fetchall()]
        intNum = [x for x in ListNum if isinstance(x, (int, float))]
        rowNum = max(intNum)
        #db.execute("""UPDATE sett set num == ?""", (rowNum, ))
        self.dialog_5_NEW.lineEdit.setText(str(int(rowNum) + 1))
        self.dialog_5_NEW.comboBox.setCurrentIndex(0)
        for i in [int(i[0]) for i in db.execute("""select  ID from Otch""").fetchall()]:
            a = db.execute("""SELECT nam from Otch WHERE ID == (?)""", (i, )).fetchone()[0]
            self.dialog_5_NEW.comboBox.addItem(str(a))
        for box_2 in "По району", "По области", "По городу":
            self.dialog_5_NEW.comboBox_2.addItem(str(box_2))
        self.dialog_5_NEW.lineEdit_18.setText('08:00')
        self.dialog_5_NEW.lineEdit_19.setText('17:00')
        self.dialog_5_NEW.lineEdit_17.setText('0')
        self.dialog_5_NEW.lineEdit_16.setText('0')
        self.dialog_5_NEW.lineEdit_21.setText('0')
        self.dialog_5_NEW.lineEdit_20.setText('0')
        self.dialog_5_NEW.lineEdit_22.setText('0')
        self.dialog_5_NEW.lineEdit_23.setText('0')
        self.dialog_5_NEW.lineEdit_24.setText('0')
        self.dialog_5_NEW.lineEdit_25.setText('0')
        self.dialog_5_NEW.lineEdit_26.setText('0')
        self.dialog_5_NEW.lineEdit_27.setText('0')
        self.dialog_5_NEW.lineEdit_8.setText('0')
        self.dialog_5_NEW.lineEdit_9.setText('0')
        self.dialog_5_NEW.lineEdit_10.setText('0')
        self.dialog_5_NEW.lineEdit_11.setText('')
        self.dialog_5_NEW.lineEdit_12.setText('')
        self.dialog_5_NEW.lineEdit_13.setText('')
        self.dialog_5_NEW.lineEdit_14.setText('')
        self.dialog_5_NEW.lineEdit_15.setText('')
        self.dialog_5_NEW.lineEdit_4.setText('00')
        self.dialog_5_NEW.lineEdit_5.setText('00')
        self.dialog_5_NEW.lineEdit_6.setText('23')
        self.dialog_5_NEW.lineEdit_7.setText('59')
        self.dialog_5_NEW.lineEdit_20.textChanged.connect(self.e_17)
        self.dialog_5_NEW.lineEdit_21.textChanged.connect(self.e_17)
        self.dialog_5_NEW.lineEdit_23.textChanged.connect(self.e_17)
        self.dialog_5_NEW.lineEdit_8.textChanged.connect(self.e_23_26)
        self.dialog_5_NEW.lineEdit_22.textChanged.connect(self.e_23_26)
        self.dialog_5_NEW.lineEdit_16.textChanged.connect(self.e_23_26)
    def e_23_26(self):
        if ((self.dialog_5_NEW.lineEdit_8.text() != '')) and (
            (self.dialog_5_NEW.lineEdit_22.text() != '')):
                  self.dialog_5_NEW.lineEdit_23.setText(str('%g' %round((((float(self.dialog_5_NEW.lineEdit_8.text()))*(
                                                              float(self.dialog_5_NEW.lineEdit_22.text()))) / 100), 1)))
        if ((self.dialog_5_NEW.lineEdit_8.text() != '')) and (
            (self.dialog_5_NEW.lineEdit_16.text() != '')):
                 self.dialog_5_NEW.lineEdit_26.setText(str('%g' % round((float(self.dialog_5_NEW.lineEdit_8.text())) + (
                                                                      float(self.dialog_5_NEW.lineEdit_16.text())), 0)))
    def e_17(self):
        if ((self.dialog_5_NEW.lineEdit_21.text() != '')) and (
            (self.dialog_5_NEW.lineEdit_20.text() != '')) and (
            (self.dialog_5_NEW.lineEdit_23.text() != '')):
               self.dialog_5_NEW.lineEdit_17.setText(str('%g' % round(((float(self.dialog_5_NEW.lineEdit_21.text())) - (
                                                                        float(self.dialog_5_NEW.lineEdit_23.text())) + (
                                                                     float(self.dialog_5_NEW.lineEdit_20.text()))), 1)))
    def but(self):
        self.dialog_5_NEW.radioButton.setChecked(True)
    def otpr(self):
        if self.dialog_5_NEW.radioButton.isChecked():
            self.dialog_5_NEW.lineEdit_23.setText(str('%g' %round((float(self.dialog_5_NEW.lineEdit_9.text())), 1)))
        else:
            self.dialog_5_NEW.lineEdit_23.setText(str('%g' %round((float(self.dialog_5_NEW.lineEdit_10.text())), 1)))
    def vr(self):
        self.dialog_5_NEW.lineEdit_4.setText('08')
        self.dialog_5_NEW.lineEdit_5.setText('00')
        self.dialog_5_NEW.lineEdit_6.setText('08')
        self.dialog_5_NEW.lineEdit_7.setText('00')
    def newline_5_NEW(self):
        try:
            Data = self.dialog_5_NEW.dateEdit.dateTime().toString('yyyy.MM.dd ddd')
            Numlist = self.dialog_5_NEW.lineEdit.text()
            Nt = self.dialog_5_NEW.lineEdit_2.text()
            Gosn = self.dialog_5_NEW.lineEdit_3.text()
            Fam = self.dialog_5_NEW.lineEdit_11.text()
            Ima = self.dialog_5_NEW.lineEdit_12.text()
            Otch = self.dialog_5_NEW.lineEdit_13.text()
            Ud = self.dialog_5_NEW.lineEdit_14.text()
            Klass = self.dialog_5_NEW.lineEdit_15.text()
            Psn = self.dialog_5_NEW.lineEdit_16.text()
            Adrp = self.dialog_5_NEW.comboBox_2.currentText()
            Vrn = self.dialog_5_NEW.lineEdit_18.text()
            Vrv = self.dialog_5_NEW.lineEdit_19.text()
            Doz = self.dialog_5_NEW.lineEdit_20.text()
            Oviezd = self.dialog_5_NEW.lineEdit_21.text()
            Ovozv = self.dialog_5_NEW.lineEdit_17.text()
            Rnorm = self.dialog_5_NEW.lineEdit_22.text()
            Rfact = self.dialog_5_NEW.lineEdit_23.text()
            Econ = self.dialog_5_NEW.lineEdit_24.text()
            Pereras = self.dialog_5_NEW.lineEdit_25.text()
            Psv = self.dialog_5_NEW.lineEdit_26.text()
            proy = self.dialog_5_NEW.lineEdit_8.text()
            ID_TRAN_1 = db.execute("""SELECT id FROM Tran WHERE (nt, gn) == (?,?)""", (Nt, Gosn,)).fetchone()[0]
            db.execute(
                """Insert into Listi ( Data, Numlist, Nt, Gosn, Fam, Ima, Otch, Ud, Klass, Psn, Adrp, Vrn, Vrv, Doz, Oviezd,
                 Ovozv, Rnorm, Rfact, Econ, Pereras, Psv, proy) values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                 (Data, Numlist, Nt, Gosn, Fam, Ima, Otch, Ud, Klass, Psn, Adrp, Vrn, Vrv, Doz, Oviezd, Ovozv, Rnorm, Rfact,
                 Econ, Pereras, Psv, proy))
            db.execute("""UPDATE Tran set speed == ? WHERE id == ?""", (Psv, ID_TRAN_1))
            db.execute("""UPDATE Tran set ost == ? WHERE id == ?""", (Ovozv, ID_TRAN_1))
            db.commit()
            self.v_vod_5()
            self.dialog_5_NEW.close()
        except:
            msg('Нужно ЗАЛОЛНИТЬ ВЫБОР ТРАНСПОРТА', 'ВНИМАНИЕ', ['ВЫБРАТЬ'])
            #self.dialog_5_NEW.close()
    def unit_5_NEW_TRAN(self):
        self.dialog_5_NEW_TRAN = uic.loadUi(resource_path("unit_5_NEW_TRAN.ui"))
        self.dialog_5_NEW_TRAN.show()
        self.dialog_5_NEW_TRAN.pushButton_2.clicked.connect(self.dialog_5_NEW_TRAN.close)
        self.dialog_5_NEW_TRAN.pushButton.clicked.connect(self.unit_5_NEW_TRAN_1)
        cur.execute('''SELECT nt, gn FROM Tran''')
        self.dialog_5_NEW_TRAN.tableWidget.setRowCount(0)
        for row, form in enumerate(cur):
            self.dialog_5_NEW_TRAN.tableWidget.insertRow(row)
            for column, item in enumerate(form):
                self.dialog_5_NEW_TRAN.tableWidget.setItem(row, column, QTableWidgetItem(str(item)))
    def unit_5_NEW_TRAN_1(self):
        i = self.dialog_5_NEW_TRAN.tableWidget.currentRow()
        nt = self.dialog_5_NEW_TRAN.tableWidget.item(i, 0).text()
        gn = self.dialog_5_NEW_TRAN.tableWidget.item(i, 1).text()
        speed = db.execute("""SELECT speed FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn, )).fetchone()[0]
        ost = db.execute("""SELECT ost FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn, )).fetchone()[0]
        norml = db.execute("""SELECT norml FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn, )).fetchone()[0]
        normz = db.execute("""SELECT normz FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn, )).fetchone()[0]
        self.dialog_5_NEW.lineEdit_2.setText(str(nt))
        self.dialog_5_NEW.lineEdit_3.setText(str(gn))
        self.dialog_5_NEW.lineEdit_16.setText(str('%g' %round((float(speed)), 1)))
        self.dialog_5_NEW.lineEdit_21.setText(str('%g' %round((float(ost)), 1)))
        if int(db.execute(""" SELECT norml From sett""").fetchone()[0]) == 1:
              self.dialog_5_NEW.lineEdit_22.setText(str('%g' %round((float(norml)), 1)))
        else: self.dialog_5_NEW.lineEdit_22.setText(str('%g' %round((float(normz)), 1)))
        self.dialog_5_NEW_TRAN.close()
    def unit_5_NEW_VOD(self):
        self.dialog_5_NEW_VOD = uic.loadUi(resource_path("unit_5_NEW_VOD.ui"))
        self.dialog_5_NEW_VOD.show()
        self.dialog_5_NEW_VOD.pushButton_2.clicked.connect(self.dialog_5_NEW_VOD.close)
        self.dialog_5_NEW_VOD.pushButton.clicked.connect(self.unit_5_NEW_VOD_1)
        cur.execute('''SELECT Fam, ima, Otch FROM vod''')
        self.dialog_5_NEW_VOD.tableWidget.setRowCount(0)
        for row, form in enumerate(cur):
            self.dialog_5_NEW_VOD.tableWidget.insertRow(row)
            for column, item in enumerate(form):
                self.dialog_5_NEW_VOD.tableWidget.setItem(row, column, QTableWidgetItem(str(item)))
    def NomID_5_VOD(self, x):
        i = self.dialog_5_NEW_VOD.tableWidget.currentRow()
        Fam = self.dialog_5_NEW_VOD.tableWidget.item(i, 0).text()
        ima = self.dialog_5_NEW_VOD.tableWidget.item(i, 1).text()
        Otch = self.dialog_5_NEW_VOD.tableWidget.item(i, 2).text()
        x = db.execute("""SELECT id FROM vod WHERE (Fam, ima, Otch) == (?, ?, ?)""", (Fam, ima, Otch,)).fetchone()[0]
        return x
    def unit_5_NEW_VOD_1(self):
        try:
            row = self.NomID_5_VOD(self.x)
            i = self.dialog_5_NEW_VOD.tableWidget.currentRow()
            Fam = self.dialog_5_NEW_VOD.tableWidget.item(i, 0).text()
            ima = self.dialog_5_NEW_VOD.tableWidget.item(i, 1).text()
            Otch = self.dialog_5_NEW_VOD.tableWidget.item(i, 2).text()
            Ud = db.execute(""" SELECT Ud From vod WHERE id == ?""", (row, )).fetchone()[0]
            klass = db.execute(""" SELECT klass From vod WHERE id == ?""", (row, )).fetchone()[0]
            self.dialog_5_NEW.lineEdit_11.setText(str(Fam))
            self.dialog_5_NEW.lineEdit_12.setText(str(ima))
            self.dialog_5_NEW.lineEdit_13.setText(str(Otch))
            self.dialog_5_NEW.lineEdit_14.setText(str(Ud))
            self.dialog_5_NEW.lineEdit_15.setText(str(klass))
            self.dialog_5_NEW_VOD.close()
        except:
            msg('Нужно выбрать строку', 'ВНИМАНИЕ', ['ВЫБРАТЬ'])
    def unit_5_RED(self):
        try:
            row = self.NomID_5(self.x)
            self.dialog_5_RED = uic.loadUi(resource_path("unit_5_NEW_RED.ui"))
            self.dialog_5_RED.show()
            self.dialog_5_RED.pushButton_7.clicked.connect(self.dialog_5_RED.close)
            self.dialog_5_RED.pushButton_6.clicked.connect(self.newline_5_RED)
            self.dialog_5_RED.pushButton_5.clicked.connect(self.unit_5_RED_VOD)
            self.dialog_5_RED.pushButton.clicked.connect(self.unit_5_RED_TRAN)
            self.but_red()
            self.e_23_26_red()
            self.e_17_red()
            self.dialog_5_RED.pushButton_2.clicked.connect(self.vr_red)
            self.dialog_5_RED.pushButton_4.clicked.connect(self.otpr_otpr)
            self.dialog_5_RED.comboBox.setCurrentIndex(0)
            for i in [int(i[0]) for i in db.execute("""select  ID from Otch""").fetchall()]:
                a = db.execute("""SELECT nam From Otch WHERE ID == ?""", (i,)).fetchone()[0]
                self.dialog_5_RED.comboBox.addItem(str(a))
            for box_2 in "По району", "По области", "По городу":
                self.dialog_5_RED.comboBox_2.addItem(str(box_2))
            self.dialog_5_RED.dateEdit.setDate(datetime.datetime.strptime((db.execute(""" SELECT Data From listi WHERE id == ?""", (row,)).fetchone()[0]), '%Y.%m.%d %a'))
            self.dialog_5_RED.dateEdit_2.setDate(datetime.datetime.strptime((db.execute(""" SELECT Data From listi WHERE id == ?""", (row,)).fetchone()[0]), '%Y.%m.%d %a'))
            self.dialog_5_RED.dateEdit_3.setDate(datetime.datetime.strptime((db.execute(""" SELECT Data From listi WHERE id == ?""", (row,)).fetchone()[0]), '%Y.%m.%d %a'))
            self.dialog_5_RED.comboBox_2.addItem(db.execute(""" SELECT adrp From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit.setText(str(db.execute(""" SELECT Numlist From listi WHERE id == ?""", (row,)).fetchone()[0]))
            self.dialog_5_RED.lineEdit_2.setText(db.execute(""" SELECT Nt From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_3.setText(db.execute(""" SELECT Gosn From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_18.setText(db.execute(""" SELECT Vrn From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_19.setText(db.execute(""" SELECT Vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_17.setText(db.execute(""" SELECT Ovozv From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_16.setText(db.execute(""" SELECT Psn From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_21.setText(db.execute(""" SELECT Oviezd From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_20.setText(db.execute(""" SELECT Doz From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_22.setText(db.execute(""" SELECT Rnorm From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_23.setText(db.execute(""" SELECT Rfact From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_24.setText(db.execute(""" SELECT Econ From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_25.setText(db.execute(""" SELECT Pereras From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_26.setText(db.execute(""" SELECT Psv From listi WHERE id == ?""", (row,)).fetchone()[0])
            # час наряда self.dialog_5_RED.lineEdit_27.setText(db.execute(""" SELECT vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_8.setText(db.execute(""" SELECT proy From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_11.setText(db.execute(""" SELECT Fam From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_12.setText(db.execute(""" SELECT Ima From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_13.setText(db.execute(""" SELECT Otch From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_14.setText(db.execute(""" SELECT Ud From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_15.setText(db.execute(""" SELECT klass From listi WHERE id == ?""", (row,)).fetchone()[0])
            self.dialog_5_RED.lineEdit_4.setText('00')
            self.dialog_5_RED.lineEdit_5.setText('00')
            self.dialog_5_RED.lineEdit_6.setText('23')
            self.dialog_5_RED.lineEdit_7.setText('59')
            self.dialog_5_RED.lineEdit_9.setText('0')
            self.dialog_5_RED.lineEdit_10.setText('0')
            self.dialog_5_RED.lineEdit_20.textChanged.connect(self.e_17_red)
            self.dialog_5_RED.lineEdit_21.textChanged.connect(self.e_17_red)
            self.dialog_5_RED.lineEdit_23.textChanged.connect(self.e_17_red)
            self.dialog_5_RED.lineEdit_8.textChanged.connect(self.e_23_26_red)
            self.dialog_5_RED.lineEdit_22.textChanged.connect(self.e_23_26_red)
            self.dialog_5_RED.lineEdit_16.textChanged.connect(self.e_23_26_red)
        except:
              msg('Нужно выбрать строку', 'ВНИМАНИЕ', ['ВЫБРАТЬ'])
              #self.dialog_5_RED.close()
    def e_23_26_red(self):
        if ((self.dialog_5_RED.lineEdit_8.text() != '')) and (
            (self.dialog_5_RED.lineEdit_22.text() != '')):
             self.dialog_5_RED.lineEdit_23.setText(str('%g' % round((((float(self.dialog_5_RED.lineEdit_8.text())) * (
                                              float(self.dialog_5_RED.lineEdit_22.text()))) / 100), 1)))
        if ((self.dialog_5_RED.lineEdit_8.text() != '')) and (
             (self.dialog_5_RED.lineEdit_16.text() != '')):
              self.dialog_5_RED.lineEdit_26.setText(str('%g' % round((float(self.dialog_5_RED.lineEdit_8.text())) + (
                                                       float(self.dialog_5_RED.lineEdit_16.text())), 0)))
    def e_17_red(self):
        if ((self.dialog_5_RED.lineEdit_21.text() != '')) and (
            (self.dialog_5_RED.lineEdit_20.text() != '')) and (
            (self.dialog_5_RED.lineEdit_23.text() != '')):
             self.dialog_5_RED.lineEdit_17.setText(str('%g' % round(((float(self.dialog_5_RED.lineEdit_21.text())) - (
            float(self.dialog_5_RED.lineEdit_23.text())) + (float(self.dialog_5_RED.lineEdit_20.text()))), 1)))
    def but_red(self):
        self.dialog_5_RED.radioButton.setChecked(True)
    def otpr_otpr(self):
        if self.dialog_5_RED.radioButton.isChecked():
            self.dialog_5_RED.lineEdit_23.setText(str('%g' %round((float(self.dialog_5_RED.lineEdit_9.text())), 1)))
        else:
            self.dialog_5_RED.lineEdit_23.setText(str('%g' %round((float(self.dialog_5_RED.lineEdit_10.text())), 1)))
    def vr_red(self):
        self.dialog_5_RED.lineEdit_4.setText('08')
        self.dialog_5_RED.lineEdit_5.setText('00')
        self.dialog_5_RED.lineEdit_6.setText('08')
        self.dialog_5_RED.lineEdit_7.setText('00')
    def newline_5_RED(self):
        row = self.NomID_5(self.x)
        Data = self.dialog_5_RED.dateEdit.dateTime().toString('yyyy.MM.dd ddd')
        Numlist = self.dialog_5_RED.lineEdit.text()
        Nt = self.dialog_5_RED.lineEdit_2.text()
        Gosn = self.dialog_5_RED.lineEdit_3.text()
        Fam = self.dialog_5_RED.lineEdit_11.text()
        Ima = self.dialog_5_RED.lineEdit_12.text()
        Otch = self.dialog_5_RED.lineEdit_13.text()
        Ud = self.dialog_5_RED.lineEdit_14.text()
        Klass = self.dialog_5_RED.lineEdit_15.text()
        Psn = self.dialog_5_RED.lineEdit_16.text()
        Adrp = self.dialog_5_RED.comboBox_2.currentText()
        Vrn = self.dialog_5_RED.lineEdit_18.text()
        Vrv = self.dialog_5_RED.lineEdit_19.text()
        Doz = self.dialog_5_RED.lineEdit_20.text()
        Oviezd = self.dialog_5_RED.lineEdit_21.text()
        Ovozv = self.dialog_5_RED.lineEdit_17.text()
        Rnorm = self.dialog_5_RED.lineEdit_22.text()
        Rfact = self.dialog_5_RED.lineEdit_23.text()
        Econ = self.dialog_5_RED.lineEdit_24.text()
        Pereras = self.dialog_5_RED.lineEdit_25.text()
        Psv = self.dialog_5_RED.lineEdit_26.text()
        proy = self.dialog_5_RED.lineEdit_8.text()
        try:
            ID_TRAN_2 = db.execute("""SELECT id FROM Tran WHERE (nt, gn) == (?,?)""", (Nt, Gosn,)).fetchone()[0]
            db.execute(
                """UPDATE listi set( Data, Nt, Gosn, Fam, Ima, Otch, Ud, Klass, Psn, Adrp, Vrn, Vrv, Doz, Oviezd,
                  Ovozv, Rnorm, Rfact, Econ, Pereras, Psv, proy) == (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) WHERE Numlist == ?""",
                  (Data, Nt, Gosn, Fam, Ima, Otch, Ud, Klass, Psn, Adrp, Vrn, Vrv, Doz, Oviezd, Ovozv, Rnorm, Rfact,
                    Econ, Pereras, Psv, proy, Numlist))
            db.execute("""UPDATE Tran set speed == ? WHERE id == ?""", (Psv, ID_TRAN_2))
            db.execute("""UPDATE Tran set ost == ? WHERE id == ?""", (Ovozv, ID_TRAN_2))
            db.commit()
            self.v_vod_5()
            self.dialog_5_RED.close()
        except:
            if ((msg('Что нужно сделать ? ', 'ВНИМАНИЕ в списке ТРАНСПОРТ нет ВЫБРАННОГО ',
                     ['ПРОДОЛЖИТЬ РЕДАКТИРОВАТЬ', 'УДАЛИТЬ ПУТЕВОЙ ЛИСТ'])) == 'УДАЛИТЬ ПУТЕВОЙ ЛИСТ'):
                row = self.NomID_5(self.x)
                db.execute("""DELETE FROM listi WHERE id == ?""", (row,))
                db.commit()
                self.v_vod_5()
                self.dialog_5_RED.close()
    def unit_5_RED_TRAN(self):
        self.dialog_5_RED_TRAN = uic.loadUi(resource_path("unit_5_RED_TRAN.ui"))
        self.dialog_5_RED_TRAN.show()
        self.dialog_5_RED_TRAN.pushButton_2.clicked.connect(self.dialog_5_RED_TRAN.close)
        self.dialog_5_RED_TRAN.pushButton.clicked.connect(self.unit_5_RED_TRAN_1)
        cur.execute('''SELECT nt, gn FROM Tran''')
        self.dialog_5_RED_TRAN.tableWidget.setRowCount(0)
        for row, form in enumerate(cur):
            self.dialog_5_RED_TRAN.tableWidget.insertRow(row)
            for column, item in enumerate(form):
                self.dialog_5_RED_TRAN.tableWidget.setItem(row, column, QTableWidgetItem(str(item)))
    def unit_5_RED_TRAN_1(self):
        i = self.dialog_5_RED_TRAN.tableWidget.currentRow()
        nt = self.dialog_5_RED_TRAN.tableWidget.item(i, 0).text()
        gn = self.dialog_5_RED_TRAN.tableWidget.item(i, 1).text()
        speed = db.execute("""SELECT speed FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn,)).fetchone()[0]
        ost = db.execute("""SELECT ost FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn,)).fetchone()[0]
        norml = db.execute("""SELECT norml FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn,)).fetchone()[0]
        normz = db.execute("""SELECT normz FROM Tran WHERE (nt, gn) == (?,?)""", (nt, gn,)).fetchone()[0]
        self.dialog_5_RED.lineEdit_2.setText(str(nt))
        self.dialog_5_RED.lineEdit_3.setText(str(gn))
        self.dialog_5_RED.lineEdit_16.setText(str('%g' % round((float(speed)), 1)))
        self.dialog_5_RED.lineEdit_21.setText(str('%g' % round((float(ost)), 1)))
        if int(db.execute(""" SELECT norml From sett""").fetchone()[0]) == 1:
            self.dialog_5_RED.lineEdit_22.setText(str('%g' % round((float(norml)), 1)))
        else:
            self.dialog_5_RED.lineEdit_22.setText(str('%g' % round((float(normz)), 1)))
        self.dialog_5_RED_TRAN.close()
    def unit_5_RED_VOD(self):
        self.dialog_5_RED_VOD = uic.loadUi(resource_path("unit_5_RED_VOD.ui"))
        self.dialog_5_RED_VOD.show()
        self.dialog_5_RED_VOD.pushButton_2.clicked.connect(self.dialog_5_RED_VOD.close)
        self.dialog_5_RED_VOD.pushButton.clicked.connect(self.unit_5_RED_VOD_1)
        cur.execute('''SELECT Fam, ima, Otch FROM vod''')
        self.dialog_5_RED_VOD.tableWidget.setRowCount(0)
        for row, form in enumerate(cur):
            self.dialog_5_RED_VOD.tableWidget.insertRow(row)
            for column, item in enumerate(form):
                self.dialog_5_RED_VOD.tableWidget.setItem(row, column, QTableWidgetItem(str(item)))
    def NomID_5_RED(self, x):
        try:
            i = self.dialog_5_RED_VOD.tableWidget.currentRow()
            Fam = self.dialog_5_RED_VOD.tableWidget.item(i, 0).text()
            ima = self.dialog_5_RED_VOD.tableWidget.item(i, 1).text()
            Otch = self.dialog_5_RED_VOD.tableWidget.item(i, 2).text()
            x = db.execute("""SELECT id FROM vod WHERE (Fam, ima, Otch) == (?, ?, ?)""", (Fam, ima, Otch,)).fetchone()[0]
            return x
        except:
            msg('Нужно выбрать строку', 'ВНИМАНИЕ', ['ВЫБРАТЬ'])
    def unit_5_RED_VOD_1(self):
        try:
            row = self.NomID_5_RED(self.x)
            i = self.dialog_5_RED_VOD.tableWidget.currentRow()
            Fam = self.dialog_5_RED_VOD.tableWidget.item(i, 0).text()
            ima = self.dialog_5_RED_VOD.tableWidget.item(i, 1).text()
            Otch = self.dialog_5_RED_VOD.tableWidget.item(i, 2).text()
            Ud = db.execute(""" SELECT Ud From vod WHERE id == ?""", (row,)).fetchone()[0]
            klass = db.execute(""" SELECT klass From vod WHERE id == ?""", (row,)).fetchone()[0]
            self.dialog_5_RED.lineEdit_11.setText(str(Fam))
            self.dialog_5_RED.lineEdit_12.setText(str(ima))
            self.dialog_5_RED.lineEdit_13.setText(str(Otch))
            self.dialog_5_RED.lineEdit_14.setText(str(Ud))
            self.dialog_5_RED.lineEdit_15.setText(str(klass))
            self.dialog_5_RED_VOD.close()
        except:
            msg('Нужно выбрать строку', 'ВНИМАНИЕ', ['ВЫБРАТЬ'])
    def NomID_5(self, x):
        i = self.dialog_5.tableWidget.currentRow()
        Data = self.dialog_5.tableWidget.item(i, 0).text()
        Numlist = self.dialog_5.tableWidget.item(i, 1).text()
        Nt = self.dialog_5.tableWidget.item(i, 2).text()
        Gosn = self.dialog_5.tableWidget.item(i, 3).text()
        Fam = self.dialog_5.tableWidget.item(i, 4).text()
        Ima = self.dialog_5.tableWidget.item(i, 5).text()
        Otch = self.dialog_5.tableWidget.item(i, 6).text()
        Ud = self.dialog_5.tableWidget.item(i, 7).text()
        Klass = self.dialog_5.tableWidget.item(i, 8).text()
        Psn = self.dialog_5.tableWidget.item(i, 9).text()
        Adrp = self.dialog_5.tableWidget.item(i, 10).text()
        Vrn = self.dialog_5.tableWidget.item(i, 11).text()
        Vrv = self.dialog_5.tableWidget.item(i, 12).text()
        Doz = self.dialog_5.tableWidget.item(i, 13).text()
        Oviezd = self.dialog_5.tableWidget.item(i, 14).text()
        Ovozv = self.dialog_5.tableWidget.item(i, 15).text()
        Rnorm = self.dialog_5.tableWidget.item(i, 16).text()
        Rfact = self.dialog_5.tableWidget.item(i, 17).text()
        Econ = self.dialog_5.tableWidget.item(i, 18).text()
        Pereras = self.dialog_5.tableWidget.item(i, 19).text()
        Psv = self.dialog_5.tableWidget.item(i, 20).text()
        proy = self.dialog_5.tableWidget.item(i, 21).text()
        x = db.execute("""SELECT id FROM listi WHERE (Data, Numlist, Nt, Gosn, Fam, Ima, Otch, Ud, Klass, Psn,
                        Adrp, Vrn, Vrv, Doz, Oviezd, Ovozv, Rnorm, Rfact, Econ, Pereras, Psv, proy) == (?, ?, ?, ?,
                         ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""", (Data, Numlist, Nt, Gosn, Fam,
                         Ima, Otch, Ud, Klass, Psn, Adrp, Vrn, Vrv, Doz, Oviezd, Ovozv, Rnorm,Rfact, Econ, Pereras,
                         Psv, proy,)).fetchone()[0]
        return x

    def unit_5_DEL(self):
        try:
            row = self.NomID_5(self.x)
            self.dialog_5_DEL = uic.loadUi(resource_path("unit_5_confirm.ui"))
            self.dialog_5_DEL.show()
            self.dialog_5_DEL.pushButton.clicked.connect(self.DEL_5)
            self.dialog_5_DEL.pushButton_2.clicked.connect(self.dialog_5_DEL.close)
        except:
            msg('Нужно выбрать строку', 'ВНИМАНИЕ', ['ВЫБРАТЬ'])
            #self.dialog_5_DEL.close()
    def DEL_5(self):
        row = self.NomID_5(self.x)
        db.execute("""DELETE FROM listi WHERE id == ?""", (row,))
        db.commit()
        self.v_vod_5()
        self.dialog_5_DEL.close()
    ######################################################################
    ######################################################################
    def unit_5_PECHAT(self):
        menu = QMenu()
        #menu.show()
        #self.dialog_5.action_4.setMenu(menu)
        menu.addAction("Легковая", self.legtr)
        menu.addAction("Грузовая", self.grutr)
        menu.addAction("Автобус").triggered.connect(self.autr)
        self.dialog_5.action_4.setMenu(menu)
    def legtr(self):
        try:
          row = self.NomID_5(self.x)
          d = datetime.datetime.strptime((db.execute(""" SELECT Data From listi WHERE id == ?""", (row,)).fetchone(
                                                                                 )[0]), '%Y.%m.%d %a')
          Adrp = (db.execute(""" SELECT adrp From listi WHERE id == ?""", (row,)).fetchone()[0])
          Numlist = (str(db.execute(""" SELECT Numlist From listi WHERE id == ?""", (row,)).fetchone()[0]))
          Nt = (db.execute(""" SELECT Nt From listi WHERE id == ?""", (row,)).fetchone()[0])
          Gosn = (db.execute(""" SELECT Gosn From listi WHERE id == ?""", (row,)).fetchone()[0])
          Vrn = (db.execute(""" SELECT Vrn From listi WHERE id == ?""", (row,)).fetchone()[0])
          Vrv = (db.execute(""" SELECT Vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
          Ovozv = (db.execute(""" SELECT Ovozv From listi WHERE id == ?""", (row,)).fetchone()[0])
          Psn = (db.execute(""" SELECT Psn From listi WHERE id == ?""", (row,)).fetchone()[0])
          Oviezd = (db.execute(""" SELECT Oviezd From listi WHERE id == ?""", (row,)).fetchone()[0])
          Doz = (db.execute(""" SELECT Doz From listi WHERE id == ?""", (row,)).fetchone()[0])
          Rnorm = (db.execute(""" SELECT Rnorm From listi WHERE id == ?""", (row,)).fetchone()[0])
          Rfact = (db.execute(""" SELECT Rfact From listi WHERE id == ?""", (row,)).fetchone()[0])
          Econ = (db.execute(""" SELECT Econ From listi WHERE id == ?""", (row,)).fetchone()[0])
          Pereras = (db.execute(""" SELECT Pereras From listi WHERE id == ?""", (row,)).fetchone()[0])
          Psv = (db.execute(""" SELECT Psv From listi WHERE id == ?""", (row,)).fetchone()[0])
         # час наряда self.dialog_5_RED.lineEdit_27.setText(db.execute(""" SELECT vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
          proy = (db.execute(""" SELECT proy From listi WHERE id == ?""", (row,)).fetchone()[0])
          Fam = (db.execute(""" SELECT Fam From listi WHERE id == ?""", (row,)).fetchone()[0])
          Ima = (db.execute(""" SELECT Ima From listi WHERE id == ?""", (row,)).fetchone()[0])
          Otch = (db.execute(""" SELECT Otch From listi WHERE id == ?""", (row,)).fetchone()[0])
          Ud = (db.execute(""" SELECT Ud From listi WHERE id == ?""", (row,)).fetchone()[0])
          klass = (db.execute(""" SELECT klass From listi WHERE id == ?""", (row,)).fetchone()[0])
          org = (db.execute(""" SELECT org From sett""").fetchone()[0])
         #self.dialog_13.lineEdit_3.setText(str(db.execute(""" SELECT adr From sett""").fetchone()[0]))
         #self.Cdialog_13.lineEdit_4.setText(str(db.execute(""" SELECT tel From sett""").fetchone()[0]))
          mex = (db.execute(""" SELECT mex From sett""").fetchone()[0])
          disp = (db.execute(""" SELECT disp From sett""").fetchone()[0])
          meds = (db.execute(""" SELECT meds From sett""").fetchone()[0])
          wb = load_workbook('legsh.xltx')
          ws1 = wb['стр1']
          ws2 = wb['стр2']
         # ws1 = wb.Worksheets(u'стр1')
         # ws2 = wb.Worksheets(u'стр2')
          ws2['E35'].value = proy
          ws1['M12'].value = ((Fam)+' '+(Ima)+' '+(Otch))
          ws1['AD5'].value = (d.day)
         ########################################
          m = ((d.month)-1)
          a = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября',
             'Декабря']
          mm = str(a[m % 12])
          ws1['AI5'].value = (mm)
         #########################################
         #########ws1['AI5'].value = (d.strftime('%B'))
         ###########################################
          ws1['AU5'].value = (d.year)
          ws1['BX4'].value = Numlist
          ws1['V10'].value = Nt
          ws1['AI11'].value = Gosn
          ws1['S14'].value = Ud
          ws1['BO14'].value = klass
          ws1['BT40'].value = Rfact
          ws1['BU19'].value = Psn
          ws1['AE30'].value = Vrn
          ws1['AG35'].value = Vrv
          ws1['BT34'].value = Doz
          ws1['BT37'].value = Oviezd
          ws1['BT38'].value = Ovozv
          ws1['R26'].value = Adrp
          ws1['BT45'].value = Psv
          ws1['BT39'].value = Rnorm
          ws1['R8'].value = org
          ws1['A22'].value = org
          ws1['AD31'].value = disp
          ws1['AD36'].value = disp
          ws1['P41'].value = meds
          ws1['BN22'].value = mex
          ws1['BN46'].value = mex
          wb.save('legsh1.xlsx')
          wb.close()
          system('xdg-open ' + 'legsh1.xlsx')
        except:
          msg('Нужно выбрать строку', 'ВНИМАНИЕ', ['ВЫБРАТЬ'])

    def grutr(self):
        pass
        # print('Грузовая')
        # row = self.NomID_5(self.x)
        # d = datetime.datetime.strptime((db.execute(""" SELECT Data From listi WHERE id == ?""", (row,)).fetchone(
        #                                                                          )[0]), '%a %d.%m.%y')
        # dv = (db.execute(""" SELECT Data From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Adrp = (db.execute(""" SELECT adrp From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Numlist = (str(db.execute(""" SELECT Numlist From listi WHERE id == ?""", (row,)).fetchone()[0]))
        # Nt = (db.execute(""" SELECT Nt From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Gosn = (db.execute(""" SELECT Gosn From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Vrn = (db.execute(""" SELECT Vrn From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Vrv = (db.execute(""" SELECT Vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Ovozv = (db.execute(""" SELECT Ovozv From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Psn = (db.execute(""" SELECT Psn From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Oviezd = (db.execute(""" SELECT Oviezd From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Doz = (db.execute(""" SELECT Doz From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Rnorm = (db.execute(""" SELECT Rnorm From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Rfact = (db.execute(""" SELECT Rfact From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Econ = (db.execute(""" SELECT Econ From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Pereras = (db.execute(""" SELECT Pereras From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Psv = (db.execute(""" SELECT Psv From listi WHERE id == ?""", (row,)).fetchone()[0])
        # # час наряда self.dialog_5_RED.lineEdit_27.setText(db.execute(""" SELECT vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
        # proy = (db.execute(""" SELECT proy From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Fam = (db.execute(""" SELECT Fam From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Ima = (db.execute(""" SELECT Ima From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Otch = (db.execute(""" SELECT Otch From listi WHERE id == ?""", (row,)).fetchone()[0])
        # Ud = (db.execute(""" SELECT Ud From listi WHERE id == ?""", (row,)).fetchone()[0])
        # klass = (db.execute(""" SELECT klass From listi WHERE id == ?""", (row,)).fetchone()[0])
        # org = (db.execute(""" SELECT org From sett""").fetchone()[0])
        # adr = (db.execute(""" SELECT adr From sett""").fetchone()[0])
        # tel = (db.execute(""" SELECT tel From sett""").fetchone()[0])
        # mex = (db.execute(""" SELECT mex From sett""").fetchone()[0])
        # disp = (db.execute(""" SELECT disp From sett""").fetchone()[0])
        # meds = (db.execute(""" SELECT meds From sett""").fetchone()[0])
        #
        # wb = load_workbook('gruzsh.xltx')
        #
        # ws1 = wb['стр1']
        # ws2 = wb['стр2']
        # # ws1 = wb.Worksheets(u'стр1')
        # # ws2 = wb.Worksheets(u'стр2')
        # ws2.Range('DI36').value = proy
        # ws2.Range('A29').value = 'Р Е З У Л Ь Т А Т Ы   Р А Б О Т Ы   А В Т О М О Б И Л Я   И   П Р И Ц Е П О В ' + 'пут.лист №' + Numlist + ' от ' + dv
        # ws2.Range('A36').value = Rnorm
        # ws2.Range('I36').value = Rfact
        # ws2.Range('AL7').value = dv
        # ws1.Range('I15').value = ((Fam)+' '+(Ima)+' '+(Otch))
        # ws1.Range('BG5').value = (d.day)
        # ########################################
        # m = ((d.month)-1)
        # a = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября',
        #      'Декабря']
        # mm = str(a[m % 12])
        # ws1.Range('BP5').value = (mm)
        # #ws1.Range('AI5').value = (d.strftime('%B'))
        # ###########################################
        #
        # ws1.Range('CL5').value = (d.year)
        # ws1.Range('CW3').value = Numlist
        # ws1.Range('Q12').value = Nt
        # ws1.Range('AB14').value = Gosn
        # ws1.Range('P17').value = Ud
        # ws1.Range('AS17').value = klass
        # ws1.Range('Q6').value = org +' '+ adr +' '+ tel
        # ws1.Range('EV14').value = Psn
        # ws1.Range('EA14').value = Vrn
        # ws1.Range('EH14').value = Vrv
        # if (d.day) < 10:
        #     mmmm = 0 + (d.day)
        #     ws1.Range('DM14').value = mmmm
        #     ws1.Range('DM15').value = mmmm
        # if (d.month) < 10:
        #     mmmm = 0 + (d.month)
        #     ws1.Range('DT14').value = mmmm
        #     ws1.Range('DT15').value = mmmm
        # ws1.Range('FH14').value = dv +' '+ Vrn
        # ws1.Range('DO24').value = Doz
        # ws1.Range('DX24').value = Oviezd
        # ws1.Range('EE24').value = Ovozv
        # ws1.Range('CE37').value = Adrp
        # ws1.Range('BT45').value = Psv
        # ws1.Range('BT39').value = Rnorm
        # ws1.Range('A37').value = org
        # ws1.Range('V46').value = disp
        # ws1.Range('AI51').value = meds
        # ws1.Range('CD43').value = mex
        # ws1.Range('CO44').value = mex
        # ws1.Range('CF54').value = mex



    def autr(self):
         pass
    #     print('Автобус')
    #     row = self.NomID_5(self.x)
    #     d = datetime.datetime.strptime((db.execute(""" SELECT Data From listi WHERE id == ?""", (row,)).fetchone(
    #                                                                              )[0]), '%a %d.%m.%y')
    #     dv = (db.execute(""" SELECT Data From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Adrp = (db.execute(""" SELECT adrp From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Numlist = (str(db.execute(""" SELECT Numlist From listi WHERE id == ?""", (row,)).fetchone()[0]))
    #     Nt = (db.execute(""" SELECT Nt From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Gosn = (db.execute(""" SELECT Gosn From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Vrn = (db.execute(""" SELECT Vrn From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Vrv = (db.execute(""" SELECT Vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Ovozv = (db.execute(""" SELECT Ovozv From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Psn = (db.execute(""" SELECT Psn From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Oviezd = (db.execute(""" SELECT Oviezd From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Doz = (db.execute(""" SELECT Doz From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Rnorm = (db.execute(""" SELECT Rnorm From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Rfact = (db.execute(""" SELECT Rfact From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Econ = (db.execute(""" SELECT Econ From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Pereras = (db.execute(""" SELECT Pereras From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Psv = (db.execute(""" SELECT Psv From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     # час наряда self.dialog_5_RED.lineEdit_27.setText(db.execute(""" SELECT vrv From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     proy = (db.execute(""" SELECT proy From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Fam = (db.execute(""" SELECT Fam From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Ima = (db.execute(""" SELECT Ima From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Otch = (db.execute(""" SELECT Otch From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     Ud = (db.execute(""" SELECT Ud From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     klass = (db.execute(""" SELECT klass From listi WHERE id == ?""", (row,)).fetchone()[0])
    #     org = (db.execute(""" SELECT org From sett""").fetchone()[0])
    #     adr = (db.execute(""" SELECT adr From sett""").fetchone()[0])
    #     tel = (db.execute(""" SELECT tel From sett""").fetchone()[0])
    #     mex = (db.execute(""" SELECT mex From sett""").fetchone()[0])
    #     disp = (db.execute(""" SELECT disp From sett""").fetchone()[0])
    #     meds = (db.execute(""" SELECT meds From sett""").fetchone()[0])
    #     #xl.Visible = True
    #
    #     wb = load_workbook(resource_path('avsh.xltx'))
    #     ws1 = wb.Worksheets(1)
    #     ws2 = wb.Worksheets(2)
    #     # ws1 = wb.Worksheets(u'стр1')
    #     # ws2 = wb.Worksheets(u'стр2')
    #     ws2.Range('FC45').value = proy
    #     ws1.Range('FD22').value = Rfact
    #     ws1.Range('P21').value = ((Fam)' '+(Ima)+' '+(Otch))
    #     ws1.Range('AF36').value = ((Fam)+' '+(Ima)+' '+(Otch))
    #     ws1.Range('BH14').value = ((Fam)+' '+(Ima)+' '+(Otch))
    #     ws1.Range('AL5').value = (d.day)
    #     ########################################
    #     m = ((d.month)-1)
    #     a = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября',
    #          'Декабря']
    #     mm = str(a[m % 12])
    #     ws1.Range('AT5').value = (mm)
    #     #ws1.Range('AI5').value = (d.strftime('%B'))
    #     ###########################################
    #
    #     ws1.Range('BK5').value = (d.year)
    #     ws1.Range('CV4').value = Numlist
    #     ws1.Range('BM6').value = Nt
    #     ws1.Range('CA7').value = Gosn
    #     ws1.Range('CS14').value = Ud
    #     ws1.Range('T43').value = Psn
    #     ws1.Range('BY33').value = Vrn
    #     ws1.Range('BK33').value = Vrv
    #     ws1.Range('FD12').value = Doz
    #     ws1.Range('FD10').value = Oviezd
    #     ws1.Range('FD15').value = Ovozv
    #     ws1.Range('T42').value = Psv
    #     ws1.Range('FD20').value = Rnorm
    #     ws1.Range('O6').value = org
    #     ws1.Range('V15').value = mex
    #     ws1.Range('AF39').value = mex
    #     ws1.Range('A9').value = adr
    #     ws1.Range('A11').value = tel
    #     ws1.Range('BR22').value = 'regn'
    #     ws1.Range('CM22').value = 'ser'
    #     ws1.Range('DA22').value = 'nomr'
    #


    def My_menu(self):
        pass

   ######################################################################
   ######################################################################

    def btnClosed(self):
        self.close()



if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MyWin()
    #window.show()
    sys.exit(app.exec_())

db.close()

